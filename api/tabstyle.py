from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side


class TabStyles:
    border_side = Side(border_style='thin', color='C5B775')
    border_style = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    font_colibri = Font(name='Calibri')
    font_colibri_bolt = Font(name='Calibri', bold=True)

    # Стиль для заголовка
    header_row_style = NamedStyle(name="header_row_style")
    header_row_style.font = font_colibri_bolt
    header_row_style.fill = PatternFill(start_color="F4ECC5", end_color="F4ECC5", fill_type="solid")
    header_row_style.border = border_style

    # Стиль для заголовка особый
    header_row_spec_style = NamedStyle(name="header_row_spec_style")
    header_row_spec_style.font = font_colibri_bolt
    header_row_spec_style.fill = PatternFill(start_color="B3AC86", end_color="B3AC86", fill_type="solid")
    header_row_spec_style.border = border_style

    # Стиль для строк первого уровня
    row_l1_style = NamedStyle(name="row_l1_style")
    row_l1_style.font = font_colibri
    row_l1_style.fill = PatternFill(start_color="FBF9EC", end_color="FBF9EC", fill_type="solid")
    row_l1_style.border = border_style

    # Стиль для строк второго уровня
    row_l2_style = NamedStyle(name="row_l2_style")
    row_l2_style.font = font_colibri
    row_l2_style.border = border_style

    # Стиль для граф рентабельности 15-й и 18-й колонки
    col_spec_style = NamedStyle(name="col_spec_style")
    col_spec_style.font = font_colibri
    col_spec_style.fill = PatternFill(start_color="FBF9EC", end_color="FBF9EC", fill_type="solid")
    col_spec_style.border = border_style

    # Стиль для строк первого уровня особый
    cell_l1_spec_style = NamedStyle(name="cell_l1_spec_style")
    cell_l1_spec_style.font = font_colibri_bolt
    cell_l1_spec_style.fill = PatternFill(start_color="B3AC86", end_color="B3AC86", fill_type="solid")
    cell_l1_spec_style.border = border_style
